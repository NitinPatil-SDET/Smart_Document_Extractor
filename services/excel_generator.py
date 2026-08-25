"""
Excel file generation from extracted data.
"""

import pandas as pd
from copy import copy
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Any, Optional
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from utils.logger import logger
from utils.helpers import ensure_directory
from utils.helpers import to_tabular_records


class ExcelGenerator:
    """
    Generate Excel files from extracted data.
    """
    
    def __init__(self, output_dir: str = "outputs"):
        """
        Initialize Excel generator.
        
        Args:
            output_dir: Directory to save Excel files
        """
        self.output_dir = ensure_directory(output_dir)
        self.logger = logger
    
    def generate_excel(self, data: Dict[str, Any], output_filename: Optional[str] = None) -> Optional[str]:
        """
        Generate Excel file from extracted data.
        
        Args:
            data: Dictionary with extracted fields
            output_filename: Name for output file (auto-generated if None)
            
        Returns:
            Path to generated Excel file or None if failed
        """
        try:
            # Generate filename if not provided
            if not output_filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_filename = f"extraction_result_{timestamp}.xlsx"
            
            output_path = self.output_dir / output_filename
            
            self.logger.info(f"Generating Excel file: {output_path}")
            
            # Expand extracted transaction columns into one worksheet row each.
            df_data = pd.DataFrame(to_tabular_records(data))
            
            # Create Excel writer
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df_data.to_excel(
                    writer,
                    sheet_name='Extracted Data',
                    index=False,
                    startrow=0
                )
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Extracted Data']
                
                # Apply formatting
                self._format_worksheet(worksheet)
            
            self.logger.info(f"Excel file created successfully: {output_path}")
            return str(output_path)
            
        except Exception as e:
            self.logger.error(f"Error generating Excel file: {e}")
            return None
    
    def generate_excel_batch(self, data_list: List[Dict[str, Any]], 
                            output_filename: Optional[str] = None) -> Optional[str]:
        """
        Generate Excel file with multiple rows of data.
        
        Args:
            data_list: List of dictionaries with extracted data
            output_filename: Name for output file (auto-generated if None)
            
        Returns:
            Path to generated Excel file or None if failed
        """
        try:
            if not data_list:
                self.logger.error("Data list is empty")
                return None
            
            # Generate filename if not provided
            if not output_filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_filename = f"extraction_batch_{timestamp}.xlsx"
            
            output_path = self.output_dir / output_filename
            
            self.logger.info(f"Generating batch Excel file: {output_path}")
            
            # Convert list of dictionaries to DataFrame
            df = pd.DataFrame(data_list)
            
            # Create Excel writer
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(
                    writer,
                    sheet_name='Extracted Data',
                    index=False,
                    startrow=0
                )
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Extracted Data']
                
                # Apply formatting
                self._format_worksheet(worksheet)
            
            self.logger.info(f"Batch Excel file created successfully: {output_path}")
            return str(output_path)
            
        except Exception as e:
            self.logger.error(f"Error generating batch Excel file: {e}")
            return None

    def get_template_columns(self, template_bytes: bytes) -> Optional[List[str]]:
        """Return column names from the first worksheet in an uploaded template."""
        try:
            workbook = load_workbook(BytesIO(template_bytes), read_only=True, data_only=False)
            worksheet = workbook[workbook.sheetnames[0]]
            columns = [cell.value for cell in worksheet[1]]
            workbook.close()
            return [str(column).strip() for column in columns if column is not None and str(column).strip()]
        except Exception as e:
            self.logger.error(f"Error reading Excel template: {e}")
            return None

    def generate_from_template(
        self,
        template_bytes: bytes,
        records: List[Dict[str, Any]],
        mapping: Dict[str, str],
        output_filename: Optional[str] = None,
    ) -> Optional[str]:
        """Populate mapped fields in the first worksheet while preserving workbook formatting."""
        try:
            if not records or not mapping:
                self.logger.error("Template records or mapping is empty")
                return None

            workbook = load_workbook(BytesIO(template_bytes))
            worksheet = workbook[workbook.sheetnames[0]]
            template_columns = {
                str(cell.value).strip(): cell.column
                for cell in worksheet[1]
                if cell.value is not None and str(cell.value).strip()
            }
            style_row = 2 if worksheet.max_row >= 2 else None

            for row_offset, record in enumerate(records, start=2):
                for template_column, extracted_field in mapping.items():
                    column_index = template_columns.get(template_column)
                    if column_index is None:
                        continue
                    target = worksheet.cell(row=row_offset, column=column_index)
                    target.value = record.get(extracted_field, "")
                    if style_row and row_offset != style_row:
                        source = worksheet.cell(row=style_row, column=column_index)
                        if source.has_style:
                            target._style = copy(source._style)
                        if source.number_format:
                            target.number_format = source.number_format
                        if source.alignment:
                            target.alignment = copy(source.alignment)

            if not output_filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_filename = f"mapped_extraction_{timestamp}.xlsx"
            output_path = self.output_dir / output_filename
            workbook.save(output_path)
            workbook.close()
            self.logger.info(f"Mapped template workbook created successfully: {output_path}")
            return str(output_path)
        except Exception as e:
            self.logger.error(f"Error populating Excel template: {e}")
            return None
    
    def _format_worksheet(self, worksheet) -> None:
        """
        Apply formatting to Excel worksheet.
        
        Args:
            worksheet: Openpyxl worksheet object
        """
        try:
            # Define header style
            header_fill = PatternFill(
                start_color="4472C4",
                end_color="4472C4",
                fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Apply header formatting
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value or "")) > max_length:
                            max_length = len(str(cell.value or ""))
                    except:
                        pass
                
                # Set column width with padding
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze header row
            worksheet.freeze_panes = "A2"
            
            # Apply center alignment to all data cells
            center_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = center_alignment
            
            self.logger.debug("Worksheet formatting applied")
            
        except Exception as e:
            self.logger.warning(f"Error applying worksheet formatting: {e}")
    
    def read_excel_data(self, excel_path: str) -> Optional[List[Dict[str, Any]]]:
        """
        Read data from Excel file.
        
        Args:
            excel_path: Path to Excel file
            
        Returns:
            List of dictionaries with data or None if failed
        """
        try:
            excel_file = Path(excel_path)
            
            if not excel_file.exists():
                self.logger.error(f"Excel file not found: {excel_path}")
                return None
            
            self.logger.info(f"Reading Excel file: {excel_path}")
            
            # Read Excel file
            df = pd.read_excel(excel_path, sheet_name='Extracted Data')
            
            # Convert to list of dictionaries
            data = df.to_dict(orient='records')
            
            self.logger.info(f"Successfully read {len(data)} rows from Excel")
            return data
            
        except Exception as e:
            self.logger.error(f"Error reading Excel file: {e}")
            return None
